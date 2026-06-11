from datetime import datetime
from io import BytesIO
from types import SimpleNamespace

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import db, Event, Gift
from config import Config


app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)


SIDE_ORDER = {
    "신부측": 1,
    "신랑측": 2,
}

RELATION_ORDER = {
    "가족": 1,
    "친척": 2,
    "친구": 3,
    "직장": 4,
    "지인": 5,
    "기타": 6,
}


def get_side_order(side):
    return SIDE_ORDER.get(side, 9)


def get_relation_order(relation):
    if not relation:
        return 999

    return RELATION_ORDER.get(relation, 100)


def sort_gifts_by_side_and_envelope(gifts):
    return sorted(
        gifts,
        key=lambda gift: (
            get_side_order(gift.side),
            gift.envelope_no or 0,
            gift.id or 0,
        ),
    )


def get_or_create_default_event():
    """
    현재는 일회용 프로그램이므로 기본 행사 1개만 사용합니다.
    events 테이블에 아무 데이터가 없으면 자동으로 하나 생성합니다.
    """
    event = Event.query.first()

    if event is None:
        event = Event(
            event_name="축의금 정리",
            default_side="신부측",
            memo="기본 행사",
        )
        db.session.add(event)
        db.session.commit()

    return event


def get_next_envelope_no(event_id, side):
    """
    신부측/신랑측 봉투번호를 각각 따로 계산합니다.

    예:
    - 신부측 #1, #2, #3
    - 신랑측 #1, #2, #3
    """
    session_key = f"next_envelope_no_{side}"

    if session.get(session_key):
        return session[session_key]

    latest_gift = (
        Gift.query
        .filter(
            Gift.event_id == event_id,
            Gift.side == side,
            Gift.deleted_at.is_(None),
        )
        .order_by(Gift.envelope_no.desc())
        .first()
    )

    if latest_gift:
        return latest_gift.envelope_no + 1

    return 1


def get_next_envelope_by_side(event_id):
    return {
        "신부측": get_next_envelope_no(event_id, "신부측"),
        "신랑측": get_next_envelope_no(event_id, "신랑측"),
    }


def is_active_envelope_no_exists(event_id, side, envelope_no):
    """
    같은 행사 안에서 같은 구분 + 같은 봉투번호가 이미 있는지 확인합니다.

    신부측 #1과 신랑측 #1은 동시에 가능하지만,
    신부측 #1이 두 번 들어가는 것은 막습니다.
    """
    return Gift.query.filter(
        Gift.event_id == event_id,
        Gift.side == side,
        Gift.envelope_no == envelope_no,
        Gift.deleted_at.is_(None),
    ).first() is not None


@app.before_request
def prepare_database():
    """
    앱 실행 후 첫 요청이 들어왔을 때 DB 테이블을 생성합니다.
    초반 개발 단계에서는 이 방식이 가장 단순합니다.
    """
    db.create_all()
    get_or_create_default_event()


@app.route("/", methods=["GET", "POST"])
def input_page():
    event = get_or_create_default_event()

    form_data = {}
    current_side = session.get("last_side", event.default_side or "신부측")

    next_envelope_by_side = get_next_envelope_by_side(event.id)
    next_envelope_no = next_envelope_by_side.get(current_side, 1)

    if request.method == "POST":
        form_data = request.form
        errors = []

        name = request.form.get("name", "").strip()
        side = request.form.get("side", "신부측").strip()
        amount_man_raw = request.form.get("amount_man", "").strip()
        meal_ticket_raw = request.form.get("meal_ticket_count", "").strip()

        relation_preset = request.form.get("relation_preset", "").strip()
        relation_custom = request.form.get("relation_custom", "").strip()
        memo = request.form.get("memo", "").strip()

        envelope_no_raw = request.form.get("envelope_no", "").strip()

        relation = relation_custom if relation_custom else relation_preset

        if side not in ["신부측", "신랑측"]:
            errors.append("구분을 다시 선택해 주세요.")
            side = "신부측"

        current_side = side

        if not name:
            errors.append("이름을 입력해 주세요.")

        envelope_no = None
        try:
            if envelope_no_raw == "":
                raise ValueError
            envelope_no = int(envelope_no_raw)
            if envelope_no <= 0:
                raise ValueError
        except ValueError:
            errors.append("봉투번호를 확인해 주세요.")

        amount_man = None
        try:
            if amount_man_raw == "":
                raise ValueError
            amount_man = int(amount_man_raw)
            if amount_man <= 0:
                raise ValueError
        except ValueError:
            errors.append("금액을 만원 단위 숫자로 입력해 주세요.")

        meal_ticket_count = None
        try:
            if meal_ticket_raw == "":
                raise ValueError
            meal_ticket_count = int(meal_ticket_raw)
            if meal_ticket_count < 0:
                raise ValueError
        except ValueError:
            errors.append("식권 수를 0 이상의 숫자로 입력해 주세요.")

        if envelope_no is not None and is_active_envelope_no_exists(event.id, side, envelope_no):
            errors.append(f"{side} #{envelope_no} 봉투번호는 이미 사용 중입니다.")

        if errors:
            for error in errors:
                flash(error, "error")

            session["last_side"] = side

            # 화면에 다시 표시할 봉투번호는 사용자가 방금 보고 있던 번호를 우선 사용합니다.
            if envelope_no is not None:
                next_envelope_no = envelope_no
            else:
                next_envelope_no = next_envelope_by_side.get(current_side, 1)

            recent_gifts = (
                Gift.query
                .filter(
                    Gift.event_id == event.id,
                    Gift.deleted_at.is_(None),
                )
                .order_by(Gift.created_at.desc(), Gift.id.desc())
                .limit(5)
                .all()
            )

            return render_template(
                "input.html",
                event=event,
                recent_gifts=recent_gifts,
                next_envelope_no=next_envelope_no,
                next_envelope_by_side=next_envelope_by_side,
                current_side=current_side,
                form_data=form_data,
            )

        amount = amount_man * 10000

        gift = Gift(
            event_id=event.id,
            envelope_no=envelope_no,
            name=name,
            side=side,
            relation=relation,
            amount=amount,
            meal_ticket_count=meal_ticket_count,
            memo=memo,
        )

        db.session.add(gift)
        db.session.commit()

        session["last_side"] = side
        session[f"next_envelope_no_{side}"] = envelope_no + 1

        flash(f"{side} #{envelope_no} {name}님 입력 완료", "success")

        return redirect(url_for("input_page"))

    recent_gifts = (
        Gift.query
        .filter(
            Gift.event_id == event.id,
            Gift.deleted_at.is_(None),
        )
        .order_by(Gift.created_at.desc(), Gift.id.desc())
        .limit(5)
        .all()
    )

    return render_template(
        "input.html",
        event=event,
        recent_gifts=recent_gifts,
        current_side=current_side,
        next_envelope_no=next_envelope_no,
        next_envelope_by_side=next_envelope_by_side,
        form_data=form_data,
    )


@app.route("/records")
def records_page():
    event = get_or_create_default_event()

    gifts = (
        Gift.query
        .filter(
            Gift.event_id == event.id,
            Gift.deleted_at.is_(None),
        )
        .all()
    )

    gifts = sort_gifts_by_side_and_envelope(gifts)

    return render_template("records.html", event=event, gifts=gifts)


@app.route("/summary")
def summary_page():
    event = get_or_create_default_event()

    gifts = (
        Gift.query
        .filter(
            Gift.event_id == event.id,
            Gift.deleted_at.is_(None),
        )
        .all()
    )

    total_amount = sum(gift.amount for gift in gifts)
    total_people = len(gifts)
    total_tickets = sum(gift.meal_ticket_count for gift in gifts)
    average_amount = total_amount // total_people if total_people > 0 else 0

    bride_gifts = [gift for gift in gifts if gift.side == "신부측"]
    groom_gifts = [gift for gift in gifts if gift.side == "신랑측"]

    bride_amount = sum(gift.amount for gift in bride_gifts)
    groom_amount = sum(gift.amount for gift in groom_gifts)

    bride_tickets = sum(gift.meal_ticket_count for gift in bride_gifts)
    groom_tickets = sum(gift.meal_ticket_count for gift in groom_gifts)

    summary = SimpleNamespace(
        total_amount=total_amount,
        total_count=total_people,
        total_tickets=total_tickets,
        average_amount=average_amount,
        bride_count=len(bride_gifts),
        groom_count=len(groom_gifts),
        bride_amount=bride_amount,
        groom_amount=groom_amount,
        bride_tickets=bride_tickets,
        groom_tickets=groom_tickets,
    )

    return render_template(
        "summary.html",
        event=event,
        summary=summary,
        total_amount=total_amount,
        total_people=total_people,
        total_tickets=total_tickets,
        average_amount=average_amount,
        bride_count=len(bride_gifts),
        groom_count=len(groom_gifts),
        bride_amount=bride_amount,
        groom_amount=groom_amount,
        bride_tickets=bride_tickets,
        groom_tickets=groom_tickets,
    )


@app.route("/gifts/<int:gift_id>/delete", methods=["POST"])
def delete_gift(gift_id):
    """
    최근 입력 카드 또는 내역 화면에서 축의금 내역을 삭제합니다.
    실제 DB row를 지우지 않고 deleted_at만 기록하는 소프트 삭제 방식입니다.

    봉투번호 매칭을 위해 삭제한 번호를 다음 입력 기본값으로 되돌립니다.
    """
    event = get_or_create_default_event()

    gift = Gift.query.filter(
        Gift.id == gift_id,
        Gift.event_id == event.id,
    ).first_or_404()

    if gift.deleted_at is None:
        gift.deleted_at = datetime.now()
        db.session.commit()

        session[f"next_envelope_no_{gift.side}"] = gift.envelope_no

        flash(f"{gift.side} #{gift.envelope_no} {gift.name}님 삭제 완료", "success")

    # 내역 화면에서 삭제했다면 내역 화면으로, 입력 화면에서 삭제했다면 입력 화면으로 돌아갑니다.
    referrer = request.referrer or ""
    if "/records" in referrer:
        return redirect(url_for("records_page"))

    return redirect(url_for("input_page"))


@app.route("/gifts/<int:gift_id>/edit", methods=["GET", "POST"])
def edit_gift(gift_id):
    event = get_or_create_default_event()

    gift = Gift.query.filter(
        Gift.id == gift_id,
        Gift.event_id == event.id,
        Gift.deleted_at.is_(None),
    ).first_or_404()

    relation_options = ["가족", "친척", "친구", "직장", "지인", "기타"]

    if request.method == "POST":
        envelope_no_raw = request.form.get("envelope_no", "").strip()
        name = request.form.get("name", "").strip()
        side = request.form.get("side", "신부측").strip()
        amount_raw = request.form.get("amount", "").strip()
        meal_ticket_raw = request.form.get("meal_ticket_count", "").strip()
        relation = request.form.get("relation", "").strip()
        relation_custom = request.form.get("relation_custom", "").strip()
        memo = request.form.get("memo", "").strip()

        if relation_custom:
            relation = relation_custom

        if not envelope_no_raw:
            flash("봉투번호를 입력해주세요.", "error")
            return render_template(
                "edit_gift.html",
                event=event,
                gift=gift,
                relation_options=relation_options,
            )

        if not name:
            flash("이름을 입력해주세요.", "error")
            return render_template(
                "edit_gift.html",
                event=event,
                gift=gift,
                relation_options=relation_options,
            )

        if side not in ["신부측", "신랑측"]:
            flash("구분을 다시 선택해주세요.", "error")
            return render_template(
                "edit_gift.html",
                event=event,
                gift=gift,
                relation_options=relation_options,
            )

        if not amount_raw:
            flash("금액을 입력해주세요.", "error")
            return render_template(
                "edit_gift.html",
                event=event,
                gift=gift,
                relation_options=relation_options,
            )

        try:
            envelope_no = int(envelope_no_raw)
            amount_man = int(amount_raw)
            meal_ticket_count = int(meal_ticket_raw or "0")
        except ValueError:
            flash("봉투번호, 금액, 식권은 숫자로 입력해주세요.", "error")
            return render_template(
                "edit_gift.html",
                event=event,
                gift=gift,
                relation_options=relation_options,
            )

        if envelope_no <= 0:
            flash("봉투번호는 1 이상이어야 합니다.", "error")
            return render_template(
                "edit_gift.html",
                event=event,
                gift=gift,
                relation_options=relation_options,
            )

        if amount_man <= 0:
            flash("금액은 1만원 이상이어야 합니다.", "error")
            return render_template(
                "edit_gift.html",
                event=event,
                gift=gift,
                relation_options=relation_options,
            )

        if meal_ticket_count < 0:
            flash("식권 수는 0 이상이어야 합니다.", "error")
            return render_template(
                "edit_gift.html",
                event=event,
                gift=gift,
                relation_options=relation_options,
            )

        duplicated_gift = Gift.query.filter(
            Gift.event_id == event.id,
            Gift.side == side,
            Gift.envelope_no == envelope_no,
            Gift.id != gift.id,
            Gift.deleted_at.is_(None),
        ).first()

        if duplicated_gift:
            flash(f"{side} #{envelope_no} 봉투번호는 이미 사용 중입니다.", "error")
            return render_template(
                "edit_gift.html",
                event=event,
                gift=gift,
                relation_options=relation_options,
            )

        gift.envelope_no = envelope_no
        gift.name = name
        gift.side = side
        gift.amount = amount_man * 10000
        gift.meal_ticket_count = meal_ticket_count
        gift.relation = relation
        gift.memo = memo
        gift.updated_at = datetime.now()

        db.session.commit()

        flash(f"{gift.side} #{gift.envelope_no} {gift.name} 내역이 수정되었습니다.", "success")
        return redirect(url_for("records_page"))

    return render_template(
        "edit_gift.html",
        event=event,
        gift=gift,
        relation_options=relation_options,
    )


@app.route("/reset-all", methods=["POST"])
def reset_all_data():
    """
    전체 입력 데이터를 초기화합니다.
    테스트 데이터 삭제 또는 행사 시작 전 리셋용입니다.

    주의:
    - gifts 테이블 전체를 비웁니다.
    - 신부측/신랑측 봉투번호를 각각 #1로 되돌립니다.
    - 구분은 신부측으로 되돌립니다.
    """
    Gift.query.delete()
    db.session.commit()

    session["next_envelope_no_신부측"] = 1
    session["next_envelope_no_신랑측"] = 1
    session["last_side"] = "신부측"

    flash("전체 데이터가 초기화되었습니다.", "success")

    return redirect(url_for("input_page"))


@app.route("/export")
def export_excel():
    """
    현재 입력된 축의금 내역을 엑셀 파일로 다운로드합니다.
    정렬은 내역 기본 화면과 같이 신부측 #1~ → 신랑측 #1~ 순서입니다.
    """
    event = get_or_create_default_event()

    gifts = (
        Gift.query
        .filter(
            Gift.event_id == event.id,
            Gift.deleted_at.is_(None),
        )
        .all()
    )
    gifts = sort_gifts_by_side_and_envelope(gifts)

    workbook = Workbook()
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    sheet = workbook.active
    sheet.title = "축의금 내역"

    # 엑셀 레이아웃
    # 1행: 제목
    # 3~4행: 필터와 연동되는 작은 요약표
    # 6행: 표 헤더
    # 7행부터: 실제 데이터
    header_row = 6
    data_start_row = header_row + 1
    data_end_row = header_row + len(gifts)
    formula_end_row = max(data_start_row, data_end_row)

    headers = [
        "구분",
        "봉투번호",
        "이름",
        "금액",
        "식권",
        "관계",
        "메모",
        "입력시간",
        "수정시간",
    ]

    # 제목
    sheet["A1"] = "축의금 내역"
    sheet.merge_cells("A1:I1")
    sheet["A1"].font = Font(bold=True, size=15, color="222222")
    sheet["A1"].alignment = Alignment(vertical="center")

    # 필터 연동 요약표
    summary_headers = ["구분", "건수", "금액", "식권"]
    summary_values = [
        "현재 표시",
        f"=SUBTOTAL(103,C{data_start_row}:C{formula_end_row})",
        f"=SUBTOTAL(109,D{data_start_row}:D{formula_end_row})",
        f"=SUBTOTAL(109,E{data_start_row}:E{formula_end_row})",
    ]

    for column_index, value in enumerate(summary_headers, start=1):
        sheet.cell(row=3, column=column_index, value=value)

    for column_index, value in enumerate(summary_values, start=1):
        sheet.cell(row=4, column=column_index, value=value)

    # 표 헤더
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=column_index, value=header)

    neutral_fill = PatternFill("solid", fgColor="EDEDED")
    summary_value_fill = PatternFill("solid", fgColor="F7F7F7")
    header_font = Font(bold=True, color="333333")
    thin_side = Side(style="thin", color="D9D9D9")
    table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 요약표 스타일
    for row_no in (3, 4):
        for column_index in range(1, 5):
            cell = sheet.cell(row=row_no, column=column_index)
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if row_no == 3:
                cell.fill = neutral_fill
                cell.font = header_font
            else:
                cell.fill = summary_value_fill

    sheet["B4"].number_format = '0"건"'
    sheet["C4"].number_format = '#,##0"원"'
    sheet["D4"].number_format = '0"장"'

    # 데이터 표 헤더 스타일
    for cell in sheet[header_row]:
        cell.fill = neutral_fill
        cell.font = header_font
        cell.border = table_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터 입력
    for row_index, gift in enumerate(gifts, start=data_start_row):
        row_values = [
            gift.side,
            gift.envelope_no,
            gift.name,
            gift.amount,
            gift.meal_ticket_count,
            gift.relation or "",
            gift.memo or "",
            gift.created_at.strftime("%Y-%m-%d %H:%M") if gift.created_at else "",
            gift.updated_at.strftime("%Y-%m-%d %H:%M") if gift.updated_at else "",
        ]

        for column_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = table_border
            cell.alignment = Alignment(vertical="center")

    column_widths = {
        "A": 10,
        "B": 9,
        "C": 14,
        "D": 13,
        "E": 7,
        "F": 10,
        "G": 24,
        "H": 18,
        "I": 18,
    }

    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width

    # 데이터 정렬 및 숫자 표시 형식
    for row in sheet.iter_rows(min_row=data_start_row, max_row=data_end_row):
        row[0].alignment = Alignment(horizontal="center", vertical="center")
        row[1].alignment = Alignment(horizontal="center", vertical="center")
        row[2].alignment = Alignment(horizontal="left", vertical="center")
        row[3].alignment = Alignment(horizontal="right", vertical="center")
        row[3].number_format = "#,##0"
        row[4].alignment = Alignment(horizontal="center", vertical="center")
        row[4].number_format = "0"
        row[5].alignment = Alignment(horizontal="left", vertical="center")
        row[6].alignment = Alignment(horizontal="left", vertical="center")
        row[7].alignment = Alignment(horizontal="center", vertical="center")
        row[8].alignment = Alignment(horizontal="center", vertical="center")

    # 필터는 표 헤더부터, 틀 고정은 데이터 시작행부터 적용합니다.
    sheet.auto_filter.ref = f"A{header_row}:I{max(header_row, data_end_row)}"
    sheet.freeze_panes = f"A{data_start_row}"

    # 입력시간/수정시간은 포함하되 기본 숨김 처리합니다.
    sheet.column_dimensions["H"].hidden = True
    sheet.column_dimensions["I"].hidden = True

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"축의금_내역_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.template_filter("money")
def money_filter(value):
    try:
        return f"{int(value):,}원"
    except (TypeError, ValueError):
        return "0원"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
